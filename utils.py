import qrcode
from io import BytesIO
import base64
from datetime import datetime
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill
from flask import make_response

def generate_qr_code(data):
    """Generate QR code and return as base64 encoded string"""
    qr = qrcode.QRCode(
        version=3,  # Increased version for better scanning
        error_correction=qrcode.constants.ERROR_CORRECT_M,  # Medium error correction
        box_size=8,  # Optimal size for scanning
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"

def generate_scannable_qr_data(code_type, product, batch, additional_data=None, base_url=""):
    """Generate structured QR code data that's easily scannable"""
    import json
    from urllib.parse import quote
    
    # Always use production URL for QR codes
    base_url = "https://translytics-uu9d.onrender.com"
    
    qr_data = {
        "type": code_type,
        "product_id": product.id,
        "product_name": product.name,
        "sku_id": product.sku_id,
        "batch_id": batch.id,
        "batch_no": batch.batch_no,
        "mfg_date": batch.mfg_date.strftime('%Y-%m-%d'),
        "expiry_date": batch.expiry_date.strftime('%Y-%m-%d'),
        "timestamp": datetime.now().strftime('%Y%m%d%H%M%S'),
        "image_url": product.image_url if product.image_url else None,
        "mrp": product.mrp if product.mrp else None,
        "gtin": product.gtin if product.gtin else None,
        "registration_no": product.registration_no if product.registration_no else None,
        "factory_name": batch.factory.name if batch.factory else None,
        "qa_status": batch.qa_status
    }
    
    if additional_data:
        qr_data.update(additional_data)
    
    # Create a comprehensive QR code that works both as URL and standalone data
    json_data = json.dumps(qr_data, separators=(',', ':'))
    encoded_data = quote(json_data)
    
    # Create a URL that points to the production server
    return f"{base_url}/scan?data={encoded_data}"

def save_uploaded_image(file, product_id):
    """Save uploaded image and return the URL"""
    import os
    from werkzeug.utils import secure_filename
    
    if file and file.filename:
        filename = secure_filename(file.filename)
        # Create unique filename with product ID
        file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'jpg'
        unique_filename = f"{product_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
        
        # Create uploads directory if it doesn't exist
        upload_dir = os.path.join(os.getcwd(), 'static', 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, unique_filename)
        file.save(file_path)
        
        # Return relative URL
        return f"/static/uploads/{unique_filename}"
    return None

def generate_batch_id():
    """Generate unique batch ID"""
    return f"BATCH{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"

def generate_product_id():
    """Generate unique product ID"""
    return f"PROD{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"

def generate_factory_id():
    """Generate unique factory ID"""
    return f"FAC{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"

def export_to_excel(data, filename, headers):
    """Export data to Excel format"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7B2CBF", end_color="7B2CBF", fill_type="solid")
    
    # Add data
    for row, item in enumerate(data, 2):
        for col, value in enumerate(item, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response
